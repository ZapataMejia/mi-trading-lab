#!/usr/bin/env python3
"""
Generador del Excel de tracking diario para los 5 bots de Polymarket.

Re-ejecutable: borra y regenera ``data/Mis_Bots_Polymarket.xlsx``.

Usage:
    python scripts/generate_tracking_excel.py
"""
from __future__ import annotations

import csv
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "data" / "Mis_Bots_Polymarket.xlsx"
BACKTEST_CSV = ROOT / "data" / "poly_backtest_year" / "may_2026_detailed.csv"

START_DATE = date(2026, 6, 3)
DAYS_AHEAD = 30  # filas 3..32 = 4 jun -> 3 jul
INITIAL_BANKROLL_PER_BOT = 100.0
NUM_BOTS = 5

BOTS = [
    ("V1 Alerts", "5pp", "Sin filtros (24/7)", 54.5, 2600, -52, 76,
     "Alto volumen, alto riesgo. Validar drawdown real."),
    ("V2B Selective", "10pp", "Skip 21/23 UTC + skip sábado + vol min $5k", 58.2, 810, -23, 12,
     "Más selectivo, mejor WR esperado."),
    ("V4A Endgame 30pp", "30pp", "Últimos 5 min del mercado", 63.8, 680, -10, 8,
     "Endgame: mucha edge en cierre. DD bajo."),
    ("V4B Endgame 40pp", "40pp", "Últimos 5 min, threshold más alto", 66.0, 304, -8, 3,
     "Versión tight. Pocos trades pero alta WR."),
    ("V4C SOL-only", "30pp", "Últimos 5 min, sólo mercados SOL", 67.5, 183, -7, 2,
     "Nicho SOL. Pocos trades. Buen edge histórico."),
]

DAY_NAMES_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

# ---------------------------------------------------------------------------
# Estilos comunes
# ---------------------------------------------------------------------------
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NAVY = "1E3A8A"
NAVY_LIGHT = "3B5BB8"
WHITE = "FFFFFF"
ZEBRA = "F4F6FA"
WEEKEND = "E8EAF1"

GREEN_STRONG = "16A34A"
GREEN_LIGHT = "BBF7D0"
YELLOW = "FEF3C7"
RED_SOFT = "FECACA"
RED_TEXT = "B91C1C"
GREEN_TEXT = "166534"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor=NAVY_LIGHT)
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=12, italic=True, color="475569")
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

USD_FMT = '"$"#,##0.00;[Red]-"$"#,##0.00'
USD_PNL_FMT = '"$"#,##0.00;[Red]-"$"#,##0.00;"$"0.00'
PCT_FMT = '0.0%'
INT_FMT = '0'


def apply_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def apply_subheader(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def apply_data(cell, *, fmt=None, align=None, zebra=False, weekend=False, bold=False):
    cell.font = Font(name="Calibri", size=11, bold=bold)
    cell.alignment = align or CENTER
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if weekend:
        cell.fill = PatternFill("solid", fgColor=WEEKEND)
    elif zebra:
        cell.fill = PatternFill("solid", fgColor=ZEBRA)


# ---------------------------------------------------------------------------
# Pestaña 2: "Día a día"  (la creamos primero porque "Resumen" la referencia)
# ---------------------------------------------------------------------------
DAILY_COLUMNS = [
    ("Fecha", 12, "date"),
    ("Día", 7, "day"),
    ("V1 Trades", 11, "int"),
    ("V1 WR %", 10, "pct"),
    ("V1 PnL día", 12, "usd"),
    ("V1 Bankroll", 13, "usd"),
    ("V2B Trades", 11, "int"),
    ("V2B WR %", 10, "pct"),
    ("V2B PnL día", 12, "usd"),
    ("V2B Bankroll", 13, "usd"),
    ("V4A Trades", 11, "int"),
    ("V4A WR %", 10, "pct"),
    ("V4A PnL día", 12, "usd"),
    ("V4A Bankroll", 13, "usd"),
    ("V4B Trades", 11, "int"),
    ("V4B WR %", 10, "pct"),
    ("V4B PnL día", 12, "usd"),
    ("V4B Bankroll", 13, "usd"),
    ("V4C Trades", 11, "int"),
    ("V4C WR %", 10, "pct"),
    ("V4C PnL día", 12, "usd"),
    ("V4C Bankroll", 13, "usd"),
    ("TOTAL PnL día", 14, "usd"),
    ("TOTAL Bankroll", 15, "usd"),
    ("Notas", 40, "text"),
]

# Mapeo de columnas por bot (1-indexed): trades, wr, pnl, bankroll
BOT_COL_GROUPS = [
    {"name": "V1",  "trades": 3,  "wr": 4,  "pnl": 5,  "bankroll": 6},
    {"name": "V2B", "trades": 7,  "wr": 8,  "pnl": 9,  "bankroll": 10},
    {"name": "V4A", "trades": 11, "wr": 12, "pnl": 13, "bankroll": 14},
    {"name": "V4B", "trades": 15, "wr": 16, "pnl": 17, "bankroll": 18},
    {"name": "V4C", "trades": 19, "wr": 20, "pnl": 21, "bankroll": 22},
]
TOTAL_PNL_COL = 23
TOTAL_BANK_COL = 24
NOTES_COL = 25

HEADER_ROW = 1
DAY_ZERO_ROW = 2
FIRST_USER_ROW = 3
LAST_USER_ROW = FIRST_USER_ROW + DAYS_AHEAD - 1  # 32
DATA_RANGE_FIRST = DAY_ZERO_ROW
DATA_RANGE_LAST = LAST_USER_ROW


def build_daily_sheet(ws):
    ws.title = "Día a día"
    ws.sheet_view.showGridLines = False

    # Header row
    ws.row_dimensions[HEADER_ROW].height = 38
    for idx, (label, width, _typ) in enumerate(DAILY_COLUMNS, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width
        c = ws.cell(row=HEADER_ROW, column=idx, value=label)
        apply_header(c)

    # ----- Row 2: Día 0 (3 jun 2026), bankrolls iniciales -----
    ws.cell(row=DAY_ZERO_ROW, column=1, value=START_DATE)
    ws.cell(row=DAY_ZERO_ROW, column=2, value=DAY_NAMES_ES[START_DATE.weekday()])
    for grp in BOT_COL_GROUPS:
        ws.cell(row=DAY_ZERO_ROW, column=grp["trades"], value=0)
        ws.cell(row=DAY_ZERO_ROW, column=grp["wr"], value=None)
        ws.cell(row=DAY_ZERO_ROW, column=grp["pnl"], value=0)
        ws.cell(row=DAY_ZERO_ROW, column=grp["bankroll"], value=INITIAL_BANKROLL_PER_BOT)
    # Totales fila 2
    total_bank_initial = INITIAL_BANKROLL_PER_BOT * NUM_BOTS
    ws.cell(row=DAY_ZERO_ROW, column=TOTAL_PNL_COL, value=0)
    ws.cell(row=DAY_ZERO_ROW, column=TOTAL_BANK_COL, value=total_bank_initial)
    ws.cell(row=DAY_ZERO_ROW, column=NOTES_COL, value="Día 0 — arranque con $100 por bot")

    # ----- Filas 3..32: fechas pre-llenadas con fórmulas -----
    for offset in range(1, DAYS_AHEAD + 1):
        row = DAY_ZERO_ROW + offset
        d = START_DATE + timedelta(days=offset)
        ws.cell(row=row, column=1, value=d)
        ws.cell(row=row, column=2, value=DAY_NAMES_ES[d.weekday()])

        prev_row = row - 1
        # Por cada bot: Bankroll = Bankroll(día anterior) + PnL(día actual, si vacío -> 0)
        for grp in BOT_COL_GROUPS:
            pnl_col = get_column_letter(grp["pnl"])
            bank_col = get_column_letter(grp["bankroll"])
            formula = f"={bank_col}{prev_row}+IF(ISBLANK({pnl_col}{row}),0,{pnl_col}{row})"
            ws.cell(row=row, column=grp["bankroll"], value=formula)

        # Totales del día
        pnl_cells = "+".join(
            f"IF(ISBLANK({get_column_letter(g['pnl'])}{row}),0,{get_column_letter(g['pnl'])}{row})"
            for g in BOT_COL_GROUPS
        )
        ws.cell(row=row, column=TOTAL_PNL_COL, value=f"={pnl_cells}")
        bank_cells = "+".join(
            f"{get_column_letter(g['bankroll'])}{row}" for g in BOT_COL_GROUPS
        )
        ws.cell(row=row, column=TOTAL_BANK_COL, value=f"={bank_cells}")

    # ----- Estilos de filas de datos -----
    for row in range(DAY_ZERO_ROW, LAST_USER_ROW + 1):
        day_name = ws.cell(row=row, column=2).value
        is_weekend = day_name in ("Sáb", "Dom")
        is_zebra = (row % 2 == 0) and not is_weekend
        for idx, (_label, _w, typ) in enumerate(DAILY_COLUMNS, start=1):
            cell = ws.cell(row=row, column=idx)
            fmt = None
            align = CENTER
            bold = idx in (TOTAL_PNL_COL, TOTAL_BANK_COL)
            if typ == "date":
                fmt = "dd/mm/yyyy"
                bold = True
            elif typ == "day":
                bold = True
            elif typ == "int":
                fmt = INT_FMT
            elif typ == "pct":
                fmt = PCT_FMT
            elif typ == "usd":
                fmt = USD_PNL_FMT
            elif typ == "text":
                align = LEFT
            apply_data(cell, fmt=fmt, align=align, zebra=is_zebra, weekend=is_weekend, bold=bold)
        ws.row_dimensions[row].height = 22

    # ----- Conditional formatting -----
    # PnL día (verde si >0, rojo si <0) y bankrolls (resaltado si por debajo de 100)
    pnl_cols = [grp["pnl"] for grp in BOT_COL_GROUPS] + [TOTAL_PNL_COL]
    bank_cols = [grp["bankroll"] for grp in BOT_COL_GROUPS] + [TOTAL_BANK_COL]
    wr_cols = [grp["wr"] for grp in BOT_COL_GROUPS]

    green_font = Font(color=GREEN_TEXT, bold=True)
    red_font = Font(color=RED_TEXT, bold=True)
    green_fill_strong = PatternFill("solid", fgColor=GREEN_LIGHT)
    yellow_fill = PatternFill("solid", fgColor=YELLOW)
    red_fill = PatternFill("solid", fgColor=RED_SOFT)
    green_fill_top = PatternFill("solid", fgColor="86EFAC")

    for col in pnl_cols:
        col_letter = get_column_letter(col)
        rng = f"{col_letter}{FIRST_USER_ROW}:{col_letter}{LAST_USER_ROW}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=["0"], font=green_font)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], font=red_font)
        )

    for col in bank_cols:
        col_letter = get_column_letter(col)
        rng = f"{col_letter}{FIRST_USER_ROW}:{col_letter}{LAST_USER_ROW}"
        if col == TOTAL_BANK_COL:
            threshold = INITIAL_BANKROLL_PER_BOT * NUM_BOTS
        else:
            threshold = INITIAL_BANKROLL_PER_BOT
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=[str(threshold)], font=green_font)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=[str(threshold)], font=red_font)
        )

    # WR coloreado por bandas
    for col in wr_cols:
        col_letter = get_column_letter(col)
        rng = f"{col_letter}{FIRST_USER_ROW}:{col_letter}{LAST_USER_ROW}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=["0.65"], fill=green_fill_top, font=Font(bold=True, color=GREEN_TEXT))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="between", formula=["0.55", "0.65"], fill=green_fill_strong, font=Font(color=GREEN_TEXT))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="between", formula=["0.45", "0.5499"], fill=yellow_fill)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0.45"], fill=red_fill, font=Font(color=RED_TEXT, bold=True))
        )

    # Fines de semana (resalto toda la fila) -> ya pintado en apply_data, pero
    # también aplicamos regla por fórmula para que sea robusto si el usuario
    # mueve filas.
    weekend_fill = PatternFill("solid", fgColor=WEEKEND)
    last_col_letter = get_column_letter(len(DAILY_COLUMNS))
    full_range = f"A{FIRST_USER_ROW}:{last_col_letter}{LAST_USER_ROW}"
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'OR($B{FIRST_USER_ROW}="Sáb",$B{FIRST_USER_ROW}="Dom")'], fill=weekend_fill),
    )

    # Freeze panes: fila 1 + columnas A y B fijas
    ws.freeze_panes = "C2"

    # Autofilter sobre el header
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{LAST_USER_ROW}"


# ---------------------------------------------------------------------------
# Pestaña 1: "Resumen"
# ---------------------------------------------------------------------------
def build_resumen_sheet(ws):
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    # Anchos
    widths = {1: 32, 2: 22, 3: 22, 4: 18, 5: 18, 6: 14, 7: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Título
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1, value="Mis Bots Polymarket — Tracking diario")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:G2")
    s = ws.cell(row=2, column=1, value="Inicio: 3 junio 2026  ·  $100 por bot  ·  $500 total inicial  ·  stake $10 por trade")
    s.font = SUBTITLE_FONT
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ---- Sección: Totales globales ----
    ws.cell(row=4, column=1, value="📊 Totales globales").font = SECTION_FONT
    ws.row_dimensions[4].height = 24

    daily = "'Día a día'"
    # Rangos relevantes
    rng_total_pnl = f"{daily}!{get_column_letter(TOTAL_PNL_COL)}{FIRST_USER_ROW}:{get_column_letter(TOTAL_PNL_COL)}{LAST_USER_ROW}"
    rng_total_bank = f"{daily}!{get_column_letter(TOTAL_BANK_COL)}{FIRST_USER_ROW}:{get_column_letter(TOTAL_BANK_COL)}{LAST_USER_ROW}"
    rng_trades_all = "+".join(
        f"SUM({daily}!{get_column_letter(grp['trades'])}{DAY_ZERO_ROW}:{get_column_letter(grp['trades'])}{LAST_USER_ROW})"
        for grp in BOT_COL_GROUPS
    )

    totals = [
        ("Días operando",
         f"=SUMPRODUCT(--({daily}!{get_column_letter(TOTAL_PNL_COL)}{FIRST_USER_ROW}:{get_column_letter(TOTAL_PNL_COL)}{LAST_USER_ROW}<>0))",
         INT_FMT),
        ("Trades totales (todos los bots)", f"={rng_trades_all}", INT_FMT),
        ("Profit acumulado total", f"=SUM({rng_total_pnl})", USD_FMT),
        ("Bankroll actual total",
         f"=IFERROR(LOOKUP(2,1/({daily}!{get_column_letter(TOTAL_BANK_COL)}{DAY_ZERO_ROW}:{get_column_letter(TOTAL_BANK_COL)}{LAST_USER_ROW}<>\"\"),"
         f"{daily}!{get_column_letter(TOTAL_BANK_COL)}{DAY_ZERO_ROW}:{get_column_letter(TOTAL_BANK_COL)}{LAST_USER_ROW}),{INITIAL_BANKROLL_PER_BOT * NUM_BOTS})",
         USD_FMT),
        ("ROI total",
         f"=IFERROR(LOOKUP(2,1/({daily}!{get_column_letter(TOTAL_BANK_COL)}{DAY_ZERO_ROW}:{get_column_letter(TOTAL_BANK_COL)}{LAST_USER_ROW}<>\"\"),"
         f"{daily}!{get_column_letter(TOTAL_BANK_COL)}{DAY_ZERO_ROW}:{get_column_letter(TOTAL_BANK_COL)}{LAST_USER_ROW})/{INITIAL_BANKROLL_PER_BOT * NUM_BOTS}-1,0)",
         PCT_FMT),
        ("Mejor día (PnL)", f"=IFERROR(MAX({rng_total_pnl}),0)", USD_FMT),
        ("Peor día (PnL)", f"=IFERROR(MIN({rng_total_pnl}),0)", USD_FMT),
        ("Días positivos", f"=COUNTIF({rng_total_pnl},\">0\")", INT_FMT),
        ("Días negativos", f"=COUNTIF({rng_total_pnl},\"<0\")", INT_FMT),
    ]

    # Headers de la tabla totales
    h1 = ws.cell(row=5, column=1, value="Métrica"); apply_header(h1)
    h2 = ws.cell(row=5, column=2, value="Valor"); apply_header(h2)
    ws.row_dimensions[5].height = 26

    for i, (label, formula, fmt) in enumerate(totals):
        r = 6 + i
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=formula)
        zebra = (i % 2 == 0)
        apply_data(lc, align=LEFT, zebra=zebra, bold=True)
        apply_data(vc, fmt=fmt, align=RIGHT, zebra=zebra, bold=True)
        ws.row_dimensions[r].height = 20

    # Colorear ROI y profit
    green_font = Font(color=GREEN_TEXT, bold=True)
    red_font = Font(color=RED_TEXT, bold=True)
    for r_metric in (8, 9, 10, 11, 12):  # profit, bankroll, roi, mejor, peor
        ws.conditional_formatting.add(
            f"B{r_metric}", CellIsRule(operator="greaterThan", formula=["0"], font=green_font)
        )
        ws.conditional_formatting.add(
            f"B{r_metric}", CellIsRule(operator="lessThan", formula=["0"], font=red_font)
        )
    # Bankroll row vs valor inicial $500
    ws.conditional_formatting.add(
        "B9", CellIsRule(operator="greaterThan", formula=["500"], font=green_font)
    )
    ws.conditional_formatting.add(
        "B9", CellIsRule(operator="lessThan", formula=["500"], font=red_font)
    )

    # ---- Sección: Comparativa por bot ----
    start_row = 6 + len(totals) + 2  # deja espacio
    ws.cell(row=start_row, column=1, value="🤖 Comparativa por bot").font = SECTION_FONT
    ws.row_dimensions[start_row].height = 24

    headers = ["Bot", "Bankroll inicial", "Bankroll actual", "Profit", "Trades", "WR ponderado", "DD máx (aprox)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row + 1, column=i, value=h)
        apply_header(c)
    ws.row_dimensions[start_row + 1].height = 28

    bot_short = ["V1 Alerts", "V2B Selective", "V4A Endgame 30pp", "V4B Endgame 40pp", "V4C SOL-only"]
    for i, grp in enumerate(BOT_COL_GROUPS):
        r = start_row + 2 + i
        trades_col = get_column_letter(grp["trades"])
        wr_col = get_column_letter(grp["wr"])
        pnl_col = get_column_letter(grp["pnl"])
        bank_col = get_column_letter(grp["bankroll"])

        rng_trades = f"{daily}!{trades_col}{FIRST_USER_ROW}:{trades_col}{LAST_USER_ROW}"
        rng_wr = f"{daily}!{wr_col}{FIRST_USER_ROW}:{wr_col}{LAST_USER_ROW}"
        rng_pnl = f"{daily}!{pnl_col}{FIRST_USER_ROW}:{pnl_col}{LAST_USER_ROW}"
        rng_bank_all = f"{daily}!{bank_col}{DAY_ZERO_ROW}:{bank_col}{LAST_USER_ROW}"

        # Bankroll actual = último no vacío
        bank_actual = (
            f"=IFERROR(LOOKUP(2,1/({rng_bank_all}<>\"\"),{rng_bank_all}),{INITIAL_BANKROLL_PER_BOT})"
        )
        # WR ponderado por trades
        wr_pond = (
            f"=IFERROR(SUMPRODUCT(IF(ISNUMBER({rng_wr}),{rng_wr},0),"
            f"IF(ISNUMBER({rng_trades}),{rng_trades},0))/"
            f"SUMPRODUCT(IF(AND(ISNUMBER({rng_wr}),ISNUMBER({rng_trades})),{rng_trades},0)),0)"
        )
        # Simpler WR pond (sin AND array que rompe en algunas versiones)
        wr_pond = (
            f"=IFERROR(SUMPRODUCT(IF(ISNUMBER({rng_wr}),{rng_wr}*{rng_trades},0))/"
            f"SUMPRODUCT(IF(ISNUMBER({rng_wr}),{rng_trades},0)),0)"
        )
        # DD aprox = MIN bankroll - bankroll inicial
        dd_aprox = f"=IFERROR(MIN({rng_bank_all})-{INITIAL_BANKROLL_PER_BOT},0)"

        cells = [
            (bot_short[i], LEFT, None),
            (INITIAL_BANKROLL_PER_BOT, RIGHT, USD_FMT),
            (bank_actual, RIGHT, USD_FMT),
            (f"=SUM({rng_pnl})", RIGHT, USD_FMT),
            (f"=SUM({rng_trades})", CENTER, INT_FMT),
            (wr_pond, CENTER, PCT_FMT),
            (dd_aprox, RIGHT, USD_FMT),
        ]
        zebra = (i % 2 == 0)
        for col_idx, (val, align, fmt) in enumerate(cells, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            apply_data(c, fmt=fmt, align=align, zebra=zebra, bold=(col_idx == 1))
        ws.row_dimensions[r].height = 22

    # Conditional formatting en la tabla de bots (Profit, DD)
    profit_range = f"D{start_row + 2}:D{start_row + 1 + len(BOT_COL_GROUPS)}"
    dd_range = f"G{start_row + 2}:G{start_row + 1 + len(BOT_COL_GROUPS)}"
    bank_actual_range = f"C{start_row + 2}:C{start_row + 1 + len(BOT_COL_GROUPS)}"
    ws.conditional_formatting.add(profit_range, CellIsRule(operator="greaterThan", formula=["0"], font=green_font))
    ws.conditional_formatting.add(profit_range, CellIsRule(operator="lessThan", formula=["0"], font=red_font))
    ws.conditional_formatting.add(dd_range, CellIsRule(operator="lessThan", formula=["0"], font=red_font))
    ws.conditional_formatting.add(bank_actual_range, CellIsRule(operator="greaterThan", formula=["100"], font=green_font))
    ws.conditional_formatting.add(bank_actual_range, CellIsRule(operator="lessThan", formula=["100"], font=red_font))

    # ---- Sección: Objetivos ----
    obj_start = start_row + 2 + len(BOT_COL_GROUPS) + 2
    ws.cell(row=obj_start, column=1, value="🎯 Objetivos").font = SECTION_FONT
    ws.row_dimensions[obj_start].height = 24

    objectives = [
        ("Mes 1 (jun)", "Validar bots en vivo — NO retirar nada", "Comparar WR/PnL real vs backtest. Ajustar filtros si hace falta."),
        ("Mes 2 (jul)", "Empezar retiros del 50% de las ganancias", "Mantener bankroll base. Retirar excedente cada domingo."),
        ("Mes 3+ (ago →)", "Escalar V4A si rinde como el backtest", "Subir stake gradualmente en el bot top performer."),
    ]
    headers_obj = ["Periodo", "Objetivo", "Detalle"]
    for i, h in enumerate(headers_obj, start=1):
        c = ws.cell(row=obj_start + 1, column=i, value=h)
        apply_header(c)
    ws.row_dimensions[obj_start + 1].height = 26

    for i, (periodo, obj, det) in enumerate(objectives):
        r = obj_start + 2 + i
        zebra = (i % 2 == 0)
        c1 = ws.cell(row=r, column=1, value=periodo)
        c2 = ws.cell(row=r, column=2, value=obj)
        c3 = ws.cell(row=r, column=3, value=det)
        apply_data(c1, align=CENTER, zebra=zebra, bold=True)
        apply_data(c2, align=LEFT, zebra=zebra, bold=True)
        apply_data(c3, align=LEFT, zebra=zebra)
        ws.row_dimensions[r].height = 32

    # Merge col 3 con dos siguientes para que las notas tengan espacio
    for i in range(len(objectives)):
        r = obj_start + 2 + i
        ws.merge_cells(start_row=r, end_row=r, start_column=3, end_column=7)


# ---------------------------------------------------------------------------
# Pestaña 3: "Bots"
# ---------------------------------------------------------------------------
def build_bots_sheet(ws):
    ws.title = "Bots"
    ws.sheet_view.showGridLines = False

    widths = {1: 22, 2: 12, 3: 38, 4: 14, 5: 18, 6: 14, 7: 14, 8: 50}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Título
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="⚙️  Configuración de los 5 bots")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    s = ws.cell(row=2, column=1, value="Cifras esperadas según backtest de mayo 2026. La realidad puede variar — usar como referencia.")
    s.font = SUBTITLE_FONT
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = [
        "Nombre", "Threshold", "Filtros",
        "WR esperado", "Profit esp. mensual",
        "DD máx esp.", "Trades/día", "Notas",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        apply_header(c)
    ws.row_dimensions[4].height = 32

    green_font = Font(color=GREEN_TEXT, bold=True)
    red_font = Font(color=RED_TEXT, bold=True)

    for i, (name, thr, fil, wr, prof, dd, tpd, note) in enumerate(BOTS):
        r = 5 + i
        zebra = (i % 2 == 0)
        cells = [
            (name, LEFT, None, True),
            (thr, CENTER, None, False),
            (fil, LEFT, None, False),
            (wr / 100.0, CENTER, PCT_FMT, False),
            (prof, RIGHT, USD_FMT, True),
            (dd / 100.0, CENTER, PCT_FMT, False),
            (tpd, CENTER, INT_FMT, False),
            (note, LEFT, None, False),
        ]
        for col_idx, (val, align, fmt, bold) in enumerate(cells, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            apply_data(c, fmt=fmt, align=align, zebra=zebra, bold=bold)
        # Profit en verde
        ws.cell(row=r, column=5).font = Font(color=GREEN_TEXT, bold=True)
        # DD en rojo (es negativo)
        ws.cell(row=r, column=6).font = Font(color=RED_TEXT, bold=True)
        ws.row_dimensions[r].height = 40


# ---------------------------------------------------------------------------
# Pestaña 4: "Backtest Mayo 2026"
# ---------------------------------------------------------------------------
def build_backtest_sheet(ws):
    ws.title = "Backtest Mayo 2026"
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:L1")
    t = ws.cell(row=1, column=1, value="🗓️  Backtest detallado — Mayo 2026")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:L2")
    s = ws.cell(row=2, column=1, value="Resultados simulados con stake $10, bankroll inicial $100 por bot, datos reales de Polymarket de mayo 2026.")
    s.font = SUBTITLE_FONT
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    if not BACKTEST_CSV.exists():
        ws.cell(row=4, column=1, value=f"CSV no encontrado: {BACKTEST_CSV}").font = Font(color="B91C1C", bold=True)
        return

    # Columnas que mostramos (filtradas / renombradas para que sea entendible)
    keep = [
        ("bot", "Bot", "text", 28),
        ("trades", "Trades", "int", 10),
        ("wr_pct", "WR %", "pct100", 10),
        ("profit_usd", "Profit USD", "usd", 14),
        ("bankroll_final", "Bankroll final", "usd", 14),
        ("roi_pct", "ROI %", "pct100", 12),
        ("best_trade", "Mejor trade", "usd", 13),
        ("worst_trade", "Peor trade", "usd", 13),
        ("dd_max_usd", "DD máx USD", "usd", 14),
        ("dd_max_pct", "DD máx %", "pct100", 12),
        ("days_positive", "Días +", "int", 10),
        ("days_negative", "Días -", "int", 10),
        ("profit_factor", "Profit Factor", "ratio", 14),
    ]

    # Headers
    header_row = 4
    for i, (_src, label, _typ, width) in enumerate(keep, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        apply_header(c)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 28

    # Leer CSV
    with BACKTEST_CSV.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    green_font = Font(color=GREEN_TEXT, bold=True)
    red_font = Font(color=RED_TEXT, bold=True)

    for i, row in enumerate(rows):
        r = header_row + 1 + i
        zebra = (i % 2 == 0)
        for col_idx, (src, _label, typ, _w) in enumerate(keep, start=1):
            raw = row.get(src, "")
            cell = ws.cell(row=r, column=col_idx)
            align = CENTER
            fmt = None
            bold = False
            if typ == "text":
                # Limpiar dobles espacios en nombres
                cell.value = " ".join(raw.split())
                align = LEFT
                bold = True
            elif typ == "int":
                try:
                    cell.value = int(float(raw))
                    fmt = INT_FMT
                except (ValueError, TypeError):
                    cell.value = raw
            elif typ == "pct100":
                try:
                    cell.value = float(raw) / 100.0
                    fmt = PCT_FMT
                except (ValueError, TypeError):
                    cell.value = raw
            elif typ == "usd":
                try:
                    cell.value = round(float(raw), 2)
                    fmt = USD_FMT
                    bold = src in ("profit_usd", "bankroll_final")
                except (ValueError, TypeError):
                    cell.value = raw
            elif typ == "ratio":
                try:
                    cell.value = round(float(raw), 2)
                    fmt = "0.00"
                except (ValueError, TypeError):
                    cell.value = raw
            apply_data(cell, fmt=fmt, align=align, zebra=zebra, bold=bold)
        ws.row_dimensions[r].height = 24

    last_row = header_row + len(rows)
    # Conditional formatting: profit verde, DD rojo
    ws.conditional_formatting.add(
        f"D{header_row+1}:D{last_row}",
        CellIsRule(operator="greaterThan", formula=["0"], font=green_font),
    )
    ws.conditional_formatting.add(
        f"D{header_row+1}:D{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], font=red_font),
    )
    ws.conditional_formatting.add(
        f"F{header_row+1}:F{last_row}",
        CellIsRule(operator="greaterThan", formula=["0"], font=green_font),
    )
    ws.conditional_formatting.add(
        f"I{header_row+1}:I{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], font=red_font),
    )
    ws.conditional_formatting.add(
        f"J{header_row+1}:J{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], font=red_font),
    )

    # Freeze panes (header + columna A)
    ws.freeze_panes = f"B{header_row+1}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    if OUTPUT.exists():
        OUTPUT.unlink()
        print(f"[clean] borrado {OUTPUT}")

    wb = Workbook()
    # openpyxl crea una sheet por defecto: la usamos para Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    ws_daily = wb.create_sheet("Día a día")
    ws_bots = wb.create_sheet("Bots")
    ws_backtest = wb.create_sheet("Backtest Mayo 2026")

    # Construir en orden: daily primero (resumen lo referencia)
    build_daily_sheet(ws_daily)
    build_resumen_sheet(ws_resumen)
    build_bots_sheet(ws_bots)
    build_backtest_sheet(ws_backtest)

    # Asegurar que la pestaña visible por defecto sea Resumen
    wb.active = wb.sheetnames.index("Resumen")

    # Color de pestañas
    wb["Resumen"].sheet_properties.tabColor = NAVY
    wb["Día a día"].sheet_properties.tabColor = "16A34A"
    wb["Bots"].sheet_properties.tabColor = "B45309"
    wb["Backtest Mayo 2026"].sheet_properties.tabColor = "6D28D9"

    wb.save(OUTPUT)
    size_kb = OUTPUT.stat().st_size / 1024
    print(f"[ok] generado {OUTPUT}  ({size_kb:.1f} KB)")
    print(f"     pestañas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
